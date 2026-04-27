import csv
from io import BytesIO, StringIO
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles

from app.grader import grade_submission
from app.llm import generate_feedback
from app.rag import get_class_content_summary, get_class_metadata, list_available_classes, search_chunks
from app.schemas import FeedbackRequest, FeedbackResponse, HealthResponse, SubmissionKind, SubmittedFile
from app.utils import combined_submission_content


BASE_DIR = Path(__file__).resolve().parent.parent
FRONTEND_DIR = BASE_DIR / "frontend"
NODE_MODULES_DIR = BASE_DIR / "node_modules"

app = FastAPI(
    title="ia-corrector-mvp",
    description="API para correccion educativa con grader objetivo y feedback generado por Ollama.",
    version="0.1.0",
)

@app.get("/", include_in_schema=False)
def playground() -> FileResponse:
    return FileResponse(FRONTEND_DIR / "index.html")


@app.get("/assets/tailwindcss", include_in_schema=False)
def frontend_tailwindcss() -> Response:
    content = (FRONTEND_DIR / "tailwindcss").read_text(encoding="utf-8")
    return Response(content=content, media_type="text/css")


@app.get("/favicon.ico", include_in_schema=False)
def favicon() -> Response:
    return Response(status_code=204)


@app.get("/assets/{asset_path:path}", include_in_schema=False)
def frontend_asset(asset_path: str) -> FileResponse:
    asset = (FRONTEND_DIR / asset_path).resolve()
    if not asset.is_file() or FRONTEND_DIR.resolve() not in asset.parents:
        raise HTTPException(status_code=404, detail="Not Found")
    return FileResponse(asset)


if NODE_MODULES_DIR.exists():
    app.mount("/node_modules", StaticFiles(directory=NODE_MODULES_DIR), name="node_modules")


@app.get("/health", response_model=HealthResponse)
def health() -> HealthResponse:
    return HealthResponse(ok=True)


@app.get("/classes")
def classes() -> list[dict[str, object]]:
    return list_available_classes()


@app.post("/feedback", response_model=FeedbackResponse)
async def feedback(payload: FeedbackRequest) -> FeedbackResponse:
    return await build_feedback_response(payload)


@app.post("/feedback/upload", response_model=FeedbackResponse)
async def feedback_upload(
    class_id: str = Form(...),
    mode: str = Form(...),
    language: str = Form(...),
    files: list[UploadFile] = File(...),
) -> FeedbackResponse:
    submitted_files = [await _upload_to_submitted_file(upload) for upload in files]
    payload = FeedbackRequest(class_id=class_id, mode=mode, language=language, files=submitted_files)
    return await build_feedback_response(payload)


async def build_feedback_response(payload: FeedbackRequest) -> FeedbackResponse:
    class_metadata = get_class_metadata(payload.class_id, payload.language)
    if class_metadata is None:
        raise HTTPException(
            status_code=404,
            detail=f"No se encontro metadata para class_id={payload.class_id} y language={payload.language}.",
        )
    class_content_summary = get_class_content_summary(payload.class_id, payload.language)

    grader_result = grade_submission(payload.language, payload.files)
    query = " ".join(
        [
            combined_submission_content(payload.files),
            " ".join(class_content_summary.get("resource_titles", [])),
            " ".join(class_content_summary.get("keywords", [])),
            " ".join(grader_result["errors"]),
            " ".join(grader_result["successes"]),
            " ".join(class_metadata.get("learning_objectives", [])),
        ]
    )
    context_chunks = search_chunks(
        class_id=payload.class_id,
        language=payload.language,
        query=query,
        top_k=4,
    )

    feedback_text = await generate_feedback(
        mode=payload.mode.value,
        language=payload.language,
        class_metadata=class_metadata,
        class_content_summary=class_content_summary,
        grader_result=grader_result,
        context_chunks=context_chunks,
    )

    return FeedbackResponse(
        ok=True,
        passed=grader_result["passed"],
        score=grader_result["score"],
        errors=grader_result["errors"],
        successes=grader_result["successes"],
        context_used=[chunk["id"] for chunk in context_chunks],
        feedback=feedback_text,
    )


async def _upload_to_submitted_file(upload: UploadFile) -> SubmittedFile:
    raw = await upload.read()
    content = _decode_uploaded_bytes(upload.filename or "archivo", upload.content_type, raw)
    return SubmittedFile(
        name=upload.filename or "archivo",
        content=content,
        mime_type=upload.content_type,
        kind=SubmissionKind.file,
    )


def _decode_uploaded_bytes(filename: str, content_type: str | None, raw: bytes) -> str:
    lower_name = filename.lower()
    lower_content_type = (content_type or "").lower()

    if lower_name.endswith(".xlsx") or "spreadsheetml" in lower_content_type:
        return _xlsx_to_text(raw)
    if lower_name.endswith(".csv") or "text/csv" in lower_content_type:
        return _csv_bytes_to_text(raw)

    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue

    return raw.decode("utf-8", errors="replace")


def _csv_bytes_to_text(raw: bytes) -> str:
    text = raw.decode("utf-8", errors="replace")
    rows = list(csv.reader(StringIO(text)))
    return "\n".join(",".join(cell for cell in row) for row in rows)


def _xlsx_to_text(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="El soporte XLSX requiere tener openpyxl instalado.")

    workbook = load_workbook(BytesIO(raw), data_only=True)
    sheet_texts: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                rows.append(",".join(values))
        if rows:
            sheet_texts.append(f"# Hoja: {sheet.title}\n" + "\n".join(rows))
    return "\n\n".join(sheet_texts)

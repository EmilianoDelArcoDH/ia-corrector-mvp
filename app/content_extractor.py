import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from app.utils import fetch_url_text, strip_html_tags


STOPWORDS = {
    "a",
    "al",
    "algo",
    "algunas",
    "algunos",
    "ante",
    "antes",
    "asi",
    "bajo",
    "cada",
    "como",
    "con",
    "contra",
    "cual",
    "de",
    "del",
    "desde",
    "donde",
    "dos",
    "el",
    "ella",
    "ellas",
    "ellos",
    "en",
    "entre",
    "era",
    "eramos",
    "eran",
    "es",
    "esa",
    "esas",
    "ese",
    "eso",
    "esta",
    "estaba",
    "estaban",
    "este",
    "esto",
    "la",
    "las",
    "le",
    "les",
    "lo",
    "los",
    "mas",
    "me",
    "mi",
    "mis",
    "mucho",
    "muy",
    "no",
    "nos",
    "o",
    "para",
    "pero",
    "por",
    "porque",
    "que",
    "se",
    "sin",
    "sobre",
    "su",
    "sus",
    "tal",
    "te",
    "tiene",
    "tienen",
    "todo",
    "un",
    "una",
    "unos",
    "unas",
    "y",
    "ya",
}


@dataclass
class ExtractedChunk:
    id: str
    class_id: str
    language: str
    title: str
    content: str
    keywords: list[str]
    resource_id: str
    resource_type: str


def load_source_manifest(path: str | Path) -> list[dict[str, Any]]:
    manifest_path = Path(path)
    with manifest_path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    if isinstance(data, dict):
        resources = data.get("resources", [])
    else:
        resources = data
    if not isinstance(resources, list):
        raise ValueError("The manifest must be a list or a dict with a 'resources' list.")
    return resources


def extract_source_text(source: dict[str, Any], base_dir: str | Path | None = None) -> str:
    source_type = _normalize_source_type(source)
    if source_type == "text":
        return str(source.get("content", "") or "").strip()
    if source_type == "url":
        url = str(source.get("url", "") or "").strip()
        return extract_text_from_url(url)
    if source_type == "file":
        path = _resolve_source_path(source, base_dir)
        return extract_text_from_file(path)
    return ""


def extract_text_from_url(url: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url)
    if parsed.scheme in {"http", "https"}:
        candidate = fetch_url_text(url)
        if candidate:
            return candidate
    return ""


def extract_text_from_file(path: str | Path) -> str:
    file_path = Path(path)
    if not file_path.exists():
        return ""

    suffix = file_path.suffix.lower()
    if suffix in {".txt", ".md", ".markdown", ".rst", ".py", ".js", ".css", ".yaml", ".yml"}:
        return _read_text(file_path)
    if suffix == ".json":
        return _read_json_text(file_path)
    if suffix in {".html", ".htm"}:
        return strip_html_tags(_read_text(file_path))
    if suffix == ".csv":
        return _read_csv_text(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_text(file_path)
    if suffix == ".pdf":
        return _read_pdf_text(file_path)
    return _read_text(file_path)


def extract_resource_record(source: dict[str, Any], base_dir: str | Path | None = None) -> dict[str, Any]:
    content = extract_source_text(source, base_dir=base_dir)
    return {
        "id": source["id"],
        "class_id": source["class_id"],
        "language": source["language"],
        "type": source.get("type", "resource"),
        "title": source.get("title", source["id"]),
        "source": source.get("source", "file"),
        "path": source.get("path"),
        "url": source.get("url"),
        "mime_type": source.get("mime_type"),
        "content": content,
        "content_length": len(content),
    }


def chunk_text(
    text: str,
    *,
    chunk_size: int = 1200,
    overlap: int = 150,
) -> list[str]:
    cleaned = _normalize_whitespace(text)
    if not cleaned:
        return []

    paragraphs = [paragraph.strip() for paragraph in re.split(r"\n{2,}", text) if paragraph.strip()]
    if not paragraphs:
        paragraphs = [cleaned]

    chunks: list[str] = []
    current = ""
    for paragraph in paragraphs:
        candidate = f"{current}\n\n{paragraph}".strip() if current else paragraph
        if len(candidate) <= chunk_size:
            current = candidate
            continue
        if current:
            chunks.append(current.strip())
        if len(paragraph) <= chunk_size:
            current = paragraph
            continue
        chunks.extend(_slice_long_text(paragraph, chunk_size=chunk_size, overlap=overlap))
        current = ""

    if current.strip():
        chunks.append(current.strip())

    if not chunks:
        chunks = _slice_long_text(cleaned, chunk_size=chunk_size, overlap=overlap)

    return [chunk.strip() for chunk in chunks if chunk.strip()]


def build_chunk_records(source: dict[str, Any], base_dir: str | Path | None = None) -> list[dict[str, Any]]:
    resource = extract_resource_record(source, base_dir=base_dir)
    chunk_size = int(source.get("chunk_size", 1200))
    overlap = int(source.get("chunk_overlap", 150))
    chunks = chunk_text(resource["content"], chunk_size=chunk_size, overlap=overlap)
    keywords = _normalize_keywords(source.get("keywords", []))
    records: list[dict[str, Any]] = []
    for index, chunk in enumerate(chunks):
        records.append(
            {
                "id": f"{resource['id']}-chunk-{index + 1:03d}",
                "class_id": resource["class_id"],
                "language": resource["language"],
                "title": f"{resource['title']} - fragmento {index + 1}",
                "content": chunk,
                "keywords": _chunk_keywords(chunk, extra_keywords=keywords),
                "resource_id": resource["id"],
                "resource_type": resource["type"],
            }
        )
    return records


def build_content_base(
    sources: list[dict[str, Any]],
    *,
    base_dir: str | Path | None = None,
) -> dict[str, Any]:
    resources = [extract_resource_record(source, base_dir=base_dir) for source in sources]
    chunks: list[dict[str, Any]] = []
    for source in sources:
        chunks.extend(build_chunk_records(source, base_dir=base_dir))
    class_catalog = _build_class_catalog(resources, chunks)

    return {
        "resources": resources,
        "chunks": chunks,
        "class_catalog": class_catalog,
    }


def _normalize_source_type(source: dict[str, Any]) -> str:
    source_type = str(source.get("source", "") or source.get("kind", "") or "").lower().strip()
    if source_type in {"url", "link", "web"}:
        return "url"
    if source_type in {"text", "inline"}:
        return "text"
    return "file"


def _resolve_source_path(source: dict[str, Any], base_dir: str | Path | None) -> Path:
    path_value = source.get("path") or source.get("file") or ""
    file_path = Path(path_value)
    if file_path.is_absolute() or base_dir is None:
        return file_path
    return Path(base_dir) / file_path


def _read_text(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def _read_csv_text(path: Path) -> str:
    raw_text = _read_text(path)
    delimiter = _detect_csv_delimiter(raw_text)
    rows = list(csv.reader(StringIO(raw_text), delimiter=delimiter))
    return "\n".join(",".join(cell for cell in row) for row in rows)


def _read_json_text(path: Path) -> str:
    try:
        data = json.loads(_read_text(path))
    except json.JSONDecodeError:
        return _read_text(path)
    return _flatten_json(data)


def _read_xlsx_text(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to extract .xlsx files.") from exc

    workbook = load_workbook(path, data_only=True)
    sheet_texts: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                rows.append(",".join(values))
        if rows:
            sheet_texts.append(f"# Hoja: {sheet.title}\n" + "\n".join(rows))
    return "\n\n".join(sheet_texts)


def _read_pdf_text(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ""

    reader = PdfReader(str(path))
    pages: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            pages.append(text)
    return "\n\n".join(pages)


def _slice_long_text(text: str, *, chunk_size: int, overlap: int) -> list[str]:
    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = min(start + chunk_size, len(text))
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(text):
            break
        start = max(0, end - overlap)
    return chunks


def _normalize_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _normalize_keywords(keywords: Any) -> list[str]:
    if isinstance(keywords, str):
        return [keywords]
    if not isinstance(keywords, list):
        return []
    return [str(keyword).strip() for keyword in keywords if str(keyword).strip()]


def _chunk_keywords(text: str, extra_keywords: list[str] | None = None, limit: int = 10) -> list[str]:
    tokens = re.findall(r"[a-zA-ZáéíóúüñÁÉÍÓÚÜÑ0-9_]+", text.lower())
    tokens = [token for token in tokens if token not in STOPWORDS and len(token) > 2]
    counts = Counter(tokens)
    keywords = [token for token, _ in counts.most_common(limit)]
    if extra_keywords:
        for keyword in extra_keywords:
            normalized = keyword.lower()
            if normalized not in keywords:
                keywords.append(normalized)
    return keywords[:limit]


def _detect_csv_delimiter(text: str) -> str:
    candidates = [",", ";", "\t", "|"]
    try:
        dialect = csv.Sniffer().sniff(text, delimiters="".join(candidates))
        if dialect.delimiter in candidates:
            return dialect.delimiter
    except csv.Error:
        pass
    counts = [(text.count(candidate), candidate) for candidate in candidates]
    counts.sort(reverse=True)
    return counts[0][1] if counts and counts[0][0] > 0 else ","


def _flatten_json(value: Any, prefix: str = "") -> str:
    parts: list[str] = []
    if isinstance(value, dict):
        for key, item in value.items():
            child_prefix = f"{prefix}{key}: " if key else prefix
            parts.append(_flatten_json(item, child_prefix))
    elif isinstance(value, list):
        for item in value:
            parts.append(_flatten_json(item, prefix))
    else:
        text = str(value).strip()
        if text:
            parts.append(f"{prefix}{text}".strip())
    return "\n".join(part for part in parts if part.strip())


def _build_class_catalog(resources: list[dict[str, Any]], chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    catalog: dict[tuple[str, str], dict[str, Any]] = {}

    for resource in resources:
        key = (resource["class_id"], resource["language"])
        entry = catalog.setdefault(
            key,
            {
                "class_id": resource["class_id"],
                "language": resource["language"],
                "resource_ids": [],
                "resource_titles": [],
                "resource_types": [],
                "keywords": [],
            },
        )
        entry["resource_ids"].append(resource["id"])
        entry["resource_titles"].append(resource["title"])
        entry["resource_types"].append(resource["type"])

    for chunk in chunks:
        key = (chunk["class_id"], chunk["language"])
        entry = catalog.setdefault(
            key,
            {
                "class_id": chunk["class_id"],
                "language": chunk["language"],
                "resource_ids": [],
                "resource_titles": [],
                "resource_types": [],
                "keywords": [],
            },
        )
        for keyword in chunk.get("keywords", []):
            if keyword not in entry["keywords"]:
                entry["keywords"].append(keyword)

    for entry in catalog.values():
        entry["resource_ids"] = list(dict.fromkeys(entry["resource_ids"]))
        entry["resource_titles"] = list(dict.fromkeys(entry["resource_titles"]))
        entry["resource_types"] = list(dict.fromkeys(entry["resource_types"]))

    return sorted(catalog.values(), key=lambda item: (item["class_id"], item["language"]))
